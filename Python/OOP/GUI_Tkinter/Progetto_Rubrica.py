import tkinter as tk
from openpyxl import Workbook, load_workbook
import os
from tkinter import messagebox
# nome del file Excel
FILE_EXCEL = "Rubrica_Contatti.xlsx"

# verifichiamo che il file esista, altrimenti creiamo con intestazione
# con le intestazioni già pronte
if not os.path.exists(FILE_EXCEL):
    # crea un nuovo file Excel vuoto (nulla tra parentesi)
    # wb = Workbook
    # Woorkbook() segnala che cre un file excel vuoto
    # se dovessi usarne uno già creato scriverei 'load_workbook()'
    wb = Workbook()
    # prende un foglio attivo nel nuovo file
    # ws = Worksheet
    ws = wb.active
    # rinomina il foglio come fra ""
    ws.title = "Contatti"
    # aggiunge la prima riga del file Excel
    # ogni stringa fra "" diventa un intestazione della tabella
    ws.append(["Nome","Cognome","Data di Nascita","Numero di telefono"])
    # salva il file Excel su disco
    wb.save(FILE_EXCEL)

    # in parole povere:
    # Se il file Excel non esiste ancora,
    # creane uno nuovo,
    # prepara il foglio Contatti,
    # scrivi le colonne della tabella
    # e salva tutto.

# dichiariamo la variabile roor
root = tk.Tk()

#dichiariamo il titolo del form
root.title("Rubrica Contatti")

#dichiariamo la dimensione
root.geometry("400x400")

# dichiariamo il colore della finestra
root.configure(
    bg="#B09A09"
)

#Dichiariamo la label Nome
tk.Label(
    root,
    text="Nome: ",
    font=("Roboto",16),
    bg="#B09A09",
    fg="#fff",
    # definiamo la posizione della label (n,w,e,s, ossia i punti cardinali)
    # di default sarebbe center
    anchor="w"
).pack(fill="x", padx=10)
# padx da la distanza dal margine della label

# creazione della entry della label nome
# inizialmente lo scrivo così 'entry = tk.Entry('
# una volta specificato il nome della variabile invece scrivo così:
entry_nome = tk.Entry(
    root,
    width=30,
    font=("Roboto",16)
)
# inizialmente lo scrivo così 'entry.pack(pady=5)'
# una volta specificato il nome della variabile invece scrivo così:
entry_nome.pack(pady=5)

# Dichiariamo la label Cognome
tk.Label(
    root,
    text="Cognome: ",
    font=("Roboto",16),
    bg="#B09A09",
    fg="#fff",
    anchor="w"
).pack(fill="x", padx=10)

# creazione della entry della label cognome
# inizialmente lo scrivo così 'entry = tk.Entry('
# una volta specificato il nome della variabile invece scrivo così:
entry_cognome= tk.Entry(
    root,
    width=30,
    font=("Roboto",16)
)
# inizialmente lo scrivo così 'entry.pack(pady=5)'
# una volta specificato il nome della variabile invece scrivo così:
entry_cognome.pack(pady=5)

# Dichiariamo la label Data di nascita
tk.Label(
    root,
    text="Data di Nascita (gg/mm/aaaa): ",
    font=("Roboto",14),
    bg="#B09A09",
    fg="#fff",
    anchor="w"
).pack(fill="x", padx=10)

# creazione della entry della label data di nascita
# inizialmente lo scrivo così 'entry = tk.Entry('
# una volta specificato il nome della variabile invece scrivo così:
entry_data = tk.Entry(
    root,
    width=30,
    font=("Roboto",14)
)
# inizialmente lo scrivo così 'entry.pack(pady=5)'
# una volta specificato il nome della variabile invece scrivo così:
entry_data.pack(pady=5)

# Dichiariamo la label Telefono
tk.Label(
    root,
    text="Telefono: ",
    font=("Roboto",12),
    bg="#B09A09",
    fg="#fff",
    anchor="w"
).pack(fill="x", padx=10)

# creazione della entry della label telefono
# inizialmente lo scrivo così 'entry = tk.Entry('
# una volta specificato il nome della variabile invece scrivo così:
entry_telefono = tk.Entry(
    root,
    width=30,
    font=("Roboto",12)
)
# inizialmente lo scrivo così 'entry.pack(pady=5)'
# una volta specificato il nome della variabile invece scrivo così:
entry_telefono.pack(pady=5)

# creo il metodo per salvare il tutto, 
# inizialmente, mentre sto creando label e entry, lascio le parentesi vuote così non sono vincolata
# def salva_dati():
#    pass
# poi:
def salva_dati():
    nome = entry_nome.get()
    cognome = entry_cognome.get()
    data = entry_data.get()
    telefono = entry_telefono.get()

    # apre un file excel esistente con il percorso 'FILE EXCEL'
    wb = load_workbook(FILE_EXCEL) 

    # prende il foglio attualmente attivo del file Excel
    ws = wb.active

    # aggiunge una nuova riga in fondo al foglio Excel
    ws.append([nome,cognome,data,telefono])

    # salva le modifiche nel file Excel
    wb.save(FILE_EXCEL)

    messaggio_svuota()
    
def messaggio_svuota():
    messagebox.showinfo(
        # titolo pop up
        "Info salvataggio"
        # msg pop up
        "Il salvataggio è andato a buon fine"
    )



    # andiamo ad indicare al programma che al salvataggio deve svuotare i campi di entry
    entry_nome.delete(0, tk.END)
    entry_cognome.delete(0, tk.END)
    entry_data.delete(0, tk.END)
    entry_telefono.delete(0, tk.END)

# - - - - - BOTTONE SALVA - - - - -

button = tk.Button(
    root,
    text="Salva contatto",
    bg="white",
    font=("Roboto", 12),
    command=salva_dati
)

button.pack(pady=20)

root.mainloop()


